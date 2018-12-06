from openpyxl import Workbook
from openpyxl import load_workbook

__author__     = ['Maryanne Wachter']
__version__    = '0.1'
__status__     = 'Development'
__date__       = 'Dec 5, 2018'

def slab_excel_to_json(wb_name='Flat_Plate_Vibration_Template.xlsx', ws_name='Slab_Properties'):
	wb = load_workbook(filename=wb_name)
	ws = wb[ws_name]
	excel_address = {'l_1': 'B3', 'l_2': 'B4', 'nu': 'B8',
					   'f_c': 'B9', 'f_y': 'B10', 'h': 'B12', 'w_c': 'B14'}
	col_address = {'c1': 'B6', 'c2': 'B7'}
	loading_address = {'sdl': 'B26', 'll_design': 'B27', 'll_vib': 'B28'}
	bay_address = {'l_1': 'D3', 'l_2': 'D4'}
	all_addresses = {'l_1': 'B3', 'l_2': 'B4', 'nu': 'B8',
					   'f_c': 'B9', 'f_y': 'B10', 'h': 'B12', 'w_c': 'B14',
					 'col_size': {'c1': 'B6', 'c2': 'B7'}, 'loading': {'sdl': 'B26', 'll_design': 'B27', 'll_vib': 'B28'}}
	slab_dict = {}
	for key, address in excel_address.items():
		slab_dict[key] = ws[address].value
	slab_dict['col_size'] = {}
	for key, address in col_address.items():
		slab_dict['col_size'][key] = ws[address].value
	slab_dict['loading'] = {}
	for key, address in loading_address.items():
		slab_dict['loading'][key] = ws[address].value
	slab_dict['bay'] = {}
	for key, address in bay_address.items():
		slab_dict['bay'][key] = ws[address].value
	return slab_dict

def batch_slab_excel_to_input(start_row, end_row=None, rho=False, wb_name=None, ws_name=None):
	wb = load_workbook(filename=wb_name + '.xlsx', data_only=True)
	ws = wb[ws_name]
	if end_row == None:
		end_row = ws.max_row
	excel_ind = {'l_1': 'B', 'l_2': 'C', 'nu': 'H', 'f_c': 'I', 'f_y': 'J', 'h': 'K', 'w_c': 'L'}
	col_ind = {'c1': 'F', 'c2': 'G'}
	loading_ind = {'sdl': 'M', 'll_design': 'N', 'll_vib': 'O'}
	bay_ind = {'l_1': 'D', 'l_2': 'E'}
	slabs = []
	for i in range(start_row, end_row + 1):
		slab_dict = {}
		for key, address in excel_ind.items():
			slab_dict[key] = ws[address + str(i)].value
		slab_dict['col_size'] = {}
		for key, address in col_ind.items():
			slab_dict['col_size'][key] = ws[address + str(i)].value
		slab_dict['loading'] = {}
		for key, address in loading_ind.items():
			slab_dict['loading'][key] = ws[address + str(i)].value
		slab_dict['bay'] = {}
		for key, address in bay_ind.items():
			slab_dict['bay'][key] = ws[address + str(i)].value
		slabs.append(slab_dict)
		if rho == True:
			# should use recursion for nested dictionaries....
			slab_dict['reinforcement'] = {}
			cf = 1.0
			mf = 1.0
			slab_dict['reinforcement'].update({'l_1': {'column': {'p': ws['P' + str(i)].value * cf, 'n1': ws['Q' + str(i)].value * cf, 'n2': ws['R' + str(i)].value * cf}, 
							   				 'middle': {'p': ws['S' + str(i)].value * mf, 'n1': ws['T' + str(i)].value * mf, 'n2': ws['U' + str(i)].value * mf}},
					   			    'l_2':  {'column': {'p': ws['V' + str(i)].value * cf, 'n1': ws['W' + str(i)].value * cf, 'n2': ws['X' + str(i)].value * cf}, 
							   				 'middle': {'p': ws['Y' + str(i)].value * mf, 'n1': ws['Z' + str(i)].value * mf, 'n2': ws['AA' + str(i)].value * mf}}, 'type': 'rho'})
			# slab_dict['reinforcement'].update({'l_1': {'column': {'p': ws['P' + str(i)].value, 'n1': ws['Q' + str(i)].value, 'n2': ws['R' + str(i)].value}, 
			# 				   				 'middle': {'p': ws['S' + str(i)].value, 'n1': ws['T' + str(i)].value, 'n2': ws['U' + str(i)].value}},
			# 		   			    'l_2':  {'column': {'p': ws['V' + str(i)].value, 'n1': ws['W' + str(i)].value, 'n2': ws['X' + str(i)].value}, 
			# 				   				 'middle': {'p': ws['Y' + str(i)].value, 'n1': ws['Z' + str(i)].value, 'n2': ws['AA' + str(i)].value}}, 'type': 'rho'})
	return slabs

def batch_slab_output_to_excel(start_row, end_row, odata, rho=True, wb_name=None, ws_name=None):
	wb= load_workbook(filename=wb_name + '.xlsx')
	try:
		ws = wb[ws_name]
	except:
		ws = wb.create_sheet(ws_name)
	if end_row == None:
		end_row = start_row + len(odata)
	excel_ind = {'l_1': 'B', 'l_2': 'C', 'nu': 'H', 'f_c': 'I', 'f_y': 'J', 'h': 'K', 'w_c': 'L', 'k_1': 'AB', 'f_i': 'AC', 'weight': 'AD'}
	col_ind = {'c1': 'F', 'c2': 'G'}
	loading_ind = {'sdl': 'M', 'll_design': 'N', 'll_vib': 'O'}
	bay_ind = {'l_1': 'D', 'l_2': 'E'}
	rho_ind  = {'l_1': {'column': {'p': 'P', 'n1': 'Q', 'n2': 'R'}, 'middle': {'p': 'S', 'n1': 'T', 'n2': 'U'}},
			    'l_2': {'column': {'p': 'V', 'n1': 'W', 'n2': 'X'}, 'middle': {'p': 'Y', 'n1': 'Z', 'n2': 'AA'}}}
	vib_ind = {'beta': 'AE', 'very_slow': 'AF', 'slow': 'AG', 'moderate': 'AH', 'fast': 'AI'}
	counter = 0
	for i in range(start_row, end_row):
		# should use recursion for nested dictionaries....
		slab_data = odata[counter]
		for key, address in excel_ind.items():
			ws[address + str(i)] = slab_data[key]
		for key, address in vib_ind.items():
			if key == 'beta':
				ws[address + str(i)] = slab_data['vib'][key]
			else:
				ws[address + str(i)] = slab_data['vib'][key]['V_13']
		# Nested info
		for key, address in col_ind.items():
			ws[address + str(i)] = slab_data[key]
		for key, address in loading_ind.items():
			ws[address + str(i)] = slab_data['loading'][key]
		for key, address in bay_ind.items():
			ws[address + str(i)] = slab_data['bay'][key]
		# Reinforcement Info
		if rho == True:
			spans = ['l_1', 'l_2']
			strip_types = ['column', 'middle']
			locations = ['p', 'n1', 'n2']
			for span in spans:
				for strip_type in strip_types:
					for location in locations:
						address = rho_ind[span][strip_type][location]
						try:
							ws[address + str(i)] = slab_data['rho'][span][strip_type][location]
						except:
							pass
		counter += 1

	wb.save(filename=wb_name + '.xlsx')


if __name__ == '__main__':
	pass
